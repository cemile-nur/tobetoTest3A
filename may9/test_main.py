
from selenium import webdriver
from selenium.webdriver.common.by import By
from time import sleep
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions 
import pytest
import openpyxl
import jsons
from constants import globalConstants

class Test_Work():

    def setup_method(self):
        #test başlangıç
        self.driver=webdriver.Chrome()
        self.driver.get(globalConstants.BASE_URL)
        self.driver.maximize_window()
        sleep(5)

    def teardown_method(self):
         #test bitiş
         self.driver.quit()

    def getDataExcel():
        excelFile=openpyxl.load_workbook(globalConstants.excel_Url)
        sheet=excelFile["Sheet1"]
        rows= sheet.max_row
        data =[]
        for i in range(2,rows+1):
            username = sheet.cell(i,1).value
            password =sheet.cell(i,2).value
            data.append((username,password))
        return data
    
    @pytest.mark.parametrize("username,password", getDataExcel())
    def test_invalid_login(self,username,password):

        usernameInput = self.driver.find_element(By.ID,globalConstants.username_id)
        usernameInput.send_keys(username)
        passwordInput =self.driver.find_element(By.ID,globalConstants.password_id)
        passwordInput.send_keys(password)
        loginButton = self.driver.find_element(By.ID,globalConstants.login_button_id)
        loginButton.click()
        errorMessage = self.driver.find_element(By.XPATH,globalConstants.errorMessage_xpath)
        assert errorMessage.text == "Epic sadface: Username and password do not match any user in this service"
    

    def test_incorrectConditions(self):
       
        sleep(5)
        username=self.driver.find_element(By.ID,globalConstants.username_id)
        username.send_keys("")
        password=self.driver.find_element(By.XPATH,"//input[@name='password']")
        password.send_keys("")
        logginButton=self.driver.find_element(By.NAME,globalConstants.login_button_id)
        logginButton.click()
        sleep(2)
        errorMassage=self.driver.find_element(By.XPATH,globalConstants.errorMessage_xpath_blank)
        assert errorMassage.text=="Epic sadface: Username is required"
       
    def getData():
        excelFile = openpyxl.load_workbook(globalConstants.successExcel) 
        sheet = excelFile["Sayfa1"] 
        rows = sheet.max_row 
        data = []
        for i in range(2,rows+1):
            username = sheet.cell(i,1).value 
            password = sheet.cell(i,2).value
            data.append((username,password))
    
        return data


    @pytest.mark.parametrize("username,password",getData())

    def test_login(self,username,password):
        username=self.driver.find_element(By.ID,globalConstants.username_id)
        username.send_keys(username)
        sleep(2)
        password=self.driver.find_element(By.ID,globalConstants.password_id)
        password.send_keys(password)
        sleep(3)
        logginButton=self.driver.find_element(By.ID,globalConstants.login_button_id)
        logginButton.click()
        sleep(3)
        title=self.driver.find_element(By.XPATH,globalConstants.title_path)
        assert title.text=="Swag Labs"

        
    def getDataJSON():
        with open("may9/data/data.json") as file:
            data =jsons.load(file)
            # "account" anahtarının altındaki kullanıcı bilgilerini al
            accounts = data.get("account", [])
             # Kullanıcı adı ve şifreleri çıkar
            usernames = [account["username"] for account in accounts]
            passwords = [account["password"] for account in accounts]
            # Kullanıcı adı ve şifreleri birleştirip zip ile tuple oluştur
            return zip(usernames, passwords)


    
    @pytest.mark.parametrize("username, password", getDataJSON())
    def test_incorrectUsername(self, username, password):
        user = self.driver.find_element(By.ID, globalConstants.username_id)
        user.send_keys(username)
        sleep(2)
        passw = self.driver.find_element(By.ID, globalConstants.password_id)
        passw.send_keys(password)
        logginButton = self.driver.find_element(By.ID,globalConstants.login_button_id)
        logginButton.click()
        sleep(3)
        errorMassage = self.driver.find_element(By.XPATH, globalConstants.errorLocked)
        assert errorMassage.text == "Epic sadface: Sorry, this user has been locked out."



         
        
       
